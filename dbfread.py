import dbf
from openpyxl import Workbook
from openpyxl import load_workbook
import time

print('Program for processing DBF. Developed by @alex_hamagan for Malomikhaylovskaya TG')
of = input('Please select file name:')

table = dbf.Table(of + '.dbf', codepage='cp437')
table.open()
start_t = time.time()
#wb = load_workbook()
wb = Workbook()
sheet = wb.active

count = 1

for item in table:
    str_find = item[11].encode('cp437').decode('cp866').replace('?', 'і')
    if 'Акумуляцiя' in str_find or 'Кошти після розмежування' in str_find:
        continue
    if item[6] > 0:
        myCell = 'A' + str(count)
        myCell2 = 'B' + str(count)
        myCell3 = 'C' + str(count)
        myCell4 = 'D' + str(count)
        #myCell5 = 'E' + str(count)
        sheet[myCell] = item[4][-41:-25]
        sheet[myCell2] = item[6] / 100
        sheet[myCell3] = item[11].encode('cp437').decode('cp866').replace('?', 'і').strip()
        sheet[myCell4] = item[18].encode('cp437').decode('cp866').replace('?', 'і').strip()
        count += 1
print("Записей в таблице:", count - 1)
print(round(time.time() - start_t, 3), 'sec.')
wb.save(of + '_result.xlsx')
input('Press Enter to continue...')
